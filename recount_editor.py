import csv
import tarfile
import openpyxl
import sys

dataset_name = sys.argv[1]

exp_file = open("test_expr.tsv")
read_tsv1 = csv.reader(exp_file, delimiter="\t")
colmetadata_file = open("test_colmetadata.tsv")
read_tsv2 = csv.reader(colmetadata_file, delimiter="\t")
metadata_file_path = 'metadata.xlsx'
sheet_name = 'metadata'

new_exp_file = "expression.tab"
gene_file = "genes.tab"
new_colmetadata_file = "observations.tab"
out_tar = "recount_processed.tar.gz"

human_conv_file = "human_ensembl.txt"

unconverted_count = 0
duplicate_count = 0

# creating conversion dicts
with open(human_conv_file, 'rt') as htable:
    human_conv_dict = {}
    for line in htable:
        if line.startswith("e"):  # removes first line
            continue
        split_line = line.rstrip().split()
        if len(split_line) == 2:  # removes lines with no gene symbol
            continue
        (key, val1, val2) = (split_line[0], split_line[1], split_line[2])  # creates dictionary entry
        human_conv_dict[key] = val1, val2  # creates dictionary

# opening files to write to
with open(new_exp_file, 'w') as exp, open(gene_file, 'w') as gene, open(new_colmetadata_file, 'w') as col:
    gene.write("gene\tgene_symbol\n")
    conversion_dict = human_conv_dict
    converted_genes = {}  # list of converted genes to prevent duplicates
    col_data = {}  # dictionary to contain column metadata information
    mean = 0

    for row in read_tsv1:
        if row[0].startswith("S"):  # adding "Gene" in front of experiment names (not included previously)
            row.insert(0, "Gene")
            exp.write('\t'.join(row))
            exp.write('\n')

        elif row[0].startswith("E"):
            period_idx = 0  # index of the period in the ensembl string
            for i in row[0]:
                if i == '.':
                    phrase = row[0][period_idx:]
                    replace = row[0].replace(phrase, '')
                    row[0] = replace  # deletes all characters after period in name
                period_idx += 1

            name = row[0]
            if row[0] in conversion_dict:  # searches for name in conversion dict
                for i in row[1:]:  # calculates mean of expression data
                    i = float(i)
                    mean += i
                mean /= (len(row) - 1)
                mean = abs(mean)

                if conversion_dict[name][0] not in converted_genes:  # if gene name is not in file
                    converted_genes[conversion_dict[name][0]] = mean, row  # adds gene to dict for later comparison
                else:  # if gene is already in gene dict
                    duplicate_count += 1  # adds to duplicate number
                    # if the gene is present, but has a higher mean expression data value, replaces old version of gene
                    if mean >= converted_genes[conversion_dict[name][0]][0]:
                        converted_genes[conversion_dict[name][0]] = mean, row
            else:
                unconverted_count += 1  # if gene passes none of the tests, increments number of uncoverted genes

    for i in converted_genes:  # writing contents of expression and gene files
        out_line = '\t'.join(converted_genes[i][1])  # gets each line of expression data from dict
        exp.write(out_line)
        exp.write('\n')

        gene_contents = (converted_genes[i][1][0], i)  # gets each line of gene data from dict
        out_line2 = '\t'.join(gene_contents)
        gene.write(out_line2)
        gene.write('\n')

    for row in read_tsv2:  # iterating through column metadata file
        if row[0] != 'project':  # if the row contains project, formats line
            idx = 0
            replace = row[-1].strip('c()')
            replace = replace.split('\",')
            for i in replace:
                replace[idx] = i.replace('\"', '').strip(' ')
                idx += 1
            row[-1] = replace
            col_data[row[0]] = row[-1]

    col.write('observations\t')
    
    # iterates through col_data dict to obtain column names for observations files
    for i in col_data:
        for j in range(len(col_data[i])):
            colon_idx = 0  # index of the colon in the string
            for k in col_data[i][j]:
                if k == ':':
                    phrase = col_data[i][j][colon_idx:]
                    replace = col_data[i][j].replace(phrase, '')  # deletes anything after the colon
                    col_name = replace + '\t'
                    col.write(col_name)
                    break
                colon_idx += 1
        col.write('\n')
        break

    for i in col_data:
        out_list = [i, '\t'.join(col_data[i])]
        out_line = '\t'.join(out_list)
        col.write(out_line)
        col.write('\n')

with open("test.txt", "r") as abstract_file:
    abstract_text = abstract_file.readline().strip().replace('"', '').lstrip("[1] ")

# editing metadata
wb = openpyxl.load_workbook(metadata_file_path)
ws = wb[sheet_name]

title = ws.cell(2, 2)
summary = ws.cell(3, 2)
dataset_type = ws.cell(4, 2)
geo_accession = ws.cell(7, 2)
contact_email = ws.cell(8, 2)
contact_institute = ws.cell(9, 2)
contact_name = ws.cell(10, 2)

# adding metadata attributes
dataset_type.value = "bulk-rnaseq"
title.value = dataset_name
summary.value = abstract_text
geo_accession.value = dataset_name
contact_email.value = "sament@som.umaryland.edu"
contact_institute.value = "University of Maryland Institute for Genome Sciences"
contact_name.value = "Seth Ament"

wb.save(metadata_file_path)

with tarfile.open(out_tar, "w:gz") as tar:
    for name in [new_exp_file, gene_file, new_colmetadata_file]:
        tar.add(name)

htable.close()
exp.close()
gene.close()

print(unconverted_count, "genes unconverted")
print(duplicate_count, "duplicates found")
